"""步驟4：驗收補齊後的 xlsx（SPEC 七項檢查）。"""
import json, sys
import pandas as pd
from openpyxl import load_workbook

SCRATCH = "/tmp/claude-0/-home-user-main/6d643d58-f39d-53f8-a740-df69060d043b/scratchpad/llmfill"
SRC = "/home/user/main/job-classification-kb/training_feedback/推薦模型訓練回饋.csv"
OUT = "/home/user/main/job-classification-kb/training_feedback/推薦模型訓練回饋_補齊.xlsx"
LEAVES = "/home/user/main/job-classification-kb/scripts/_dutynm_leaves.json"

DUTY = [f"duty{i}" for i in range(5)]
NEG = [f'負向推薦{i}' for i in range(1, 6)]
POS = [f'正向推薦{i}' for i in range(1, 6)]


def main():
    src = pd.read_csv(SRC, skiprows=1, dtype=str).fillna('')
    wb = load_workbook(OUT)
    ws = wb.active
    cols = [c.value for c in ws[1]]
    data = [[(c.value or '') for c in r] for r in ws.iter_rows(min_row=2)]
    out = pd.DataFrame(data, columns=cols)
    coloured = {(r - 2, cols[c - 1])
                for r in range(2, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
                if ws.cell(row=r, column=c).fill.fgColor.rgb in ('00FFF2CC', 'FFFFF2CC')}
    leaves = set(json.load(open(LEAVES, encoding='utf-8')))
    ok = True

    # 1 原資料完整性
    bad = 0
    for c in ['職缺名稱'] + DUTY + NEG + POS + ['優化緣由']:
        for i in range(len(src)):
            if src.at[i, c] and src.at[i, c] != out.at[i, c]:
                bad += 1
    print(f"[1] 原有非空欄位被更動 {bad} 格", "✅" if bad == 0 else "❌"); ok &= bad == 0

    # 2 補入格 == 上色格
    added = {(i, c) for i in range(len(src)) for c in POS
             if not src.at[i, c] and out.at[i, c]}
    print(f"[2] 補入 {len(added):,} 格｜上色 {len(coloured):,} 格｜集合相符",
          "✅" if added == coloured else "❌"); ok &= added == coloured

    # 3 補入項存在於 tCodeDutyNM
    n = sum(1 for i, c in added if out.at[i, c] not in leaves)
    print(f"[3] 補入項不存在於葉清單 {n} 筆", "✅" if n == 0 else "❌"); ok &= n == 0

    # 4/5 與負向交集、與既有正向重複
    n4 = n5 = 0
    for i, c in added:
        v = out.at[i, c]
        if v in {out.at[i, x] for x in NEG if out.at[i, x]}:
            n4 += 1
        if v in {src.at[i, x] for x in POS if src.at[i, x]}:
            n5 += 1
    print(f"[4] 與同列負向交集 {n4} 筆", "✅" if n4 == 0 else "❌"); ok &= n4 == 0
    print(f"[5] 與既有正向重複 {n5} 筆", "✅" if n5 == 0 else "❌"); ok &= n5 == 0

    # 同列重複總檢查
    dup = sum(1 for i in range(len(out))
              if len(vs := [out.at[i, c] for c in POS if out.at[i, c]]) != len(set(vs)))
    print(f"[5b] 同列正向自身重複 {dup} 筆", "✅" if dup == 0 else "❌"); ok &= dup == 0

    # 6 仍未滿5格
    short = sum(1 for i in range(len(out)) if sum(1 for c in POS if out.at[i, c]) < 5)
    print(f"[6] 補齊後仍未滿5格 {short:,} 列（允許，門檻60以下不硬補）")

    # 7 抽查
    print("\n[7] 抽查 5 列：")
    for i in sorted({i for i, _ in added})[:5]:
        print(f"  {out.at[i,'職缺名稱'][:30]}｜{out.at[i,'優化緣由']}")
        print(f"    廠商duty : {[src.at[i,c] for c in DUTY if src.at[i,c]]}")
        print(f"    負向     : {[out.at[i,c] for c in NEG if out.at[i,c]]}")
        print(f"    正向(★=補): {[('★'+out.at[i,c]) if (i,c) in added else out.at[i,c] for c in POS if out.at[i,c]]}")

    print("\n驗收", "全數通過 ✅" if ok else "未通過 ❌")
    return 0 if ok else 1


if __name__ == '__main__':
    sys.exit(main())
