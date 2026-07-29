"""步驟3：套用 LLM 候選補齊正向推薦空格，產出 xlsx 並將補入格標底色。

規則（依 SPEC_正向推薦LLM補齊.md）：
- 只補空格，既有資料一格不動
- 候選依信心分數高→低，需通過四道過濾：存在於 tCodeDutyNM／不與既有正向重複／
  不在該列負向推薦／同列未重複
- 信心分數門檻 60（業務端確認）
- 補入格套淡黃底色 FFF2CC
"""
import json, glob, sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

SCRATCH = "/tmp/claude-0/-home-user-main/6d643d58-f39d-53f8-a740-df69060d043b/scratchpad/llmfill"
SRC = "/home/user/main/job-classification-kb/training_feedback/推薦模型訓練回饋.csv"
OUT = "/home/user/main/job-classification-kb/training_feedback/推薦模型訓練回饋_補齊.xlsx"
LEAVES = "/home/user/main/job-classification-kb/scripts/_dutynm_leaves.json"
THRESHOLD = 60
FILL = PatternFill('solid', fgColor='FFF2CC')

DUTY = [f"duty{i}" for i in range(5)]
NEG = [f'負向推薦{i}' for i in range(1, 6)]
POS = [f'正向推薦{i}' for i in range(1, 6)]


def main():
    leaves = set(json.load(open(LEAVES, encoding='utf-8')))
    cand = {}
    for f in glob.glob(f"{SCRATCH}/cache/b*.json"):
        cand.update(json.load(open(f, encoding='utf-8')))
    print(f"候選快取涵蓋標題 {len(cand):,}")

    df = pd.read_csv(SRC, skiprows=1, dtype=str).fillna('')
    cols = ['職缺名稱'] + DUTY + NEG + POS + ['優化緣由']
    df = df[cols]

    filled_cells = set()      # (row_idx, col_name)
    stat = {'rows_filled': 0, 'cells_filled': 0, 'rows_still_short': 0,
            'rej_notleaf': 0, 'rej_dup': 0, 'rej_neg': 0, 'rej_score': 0}

    for ri, row in df.iterrows():
        gaps = [c for c in POS if not row[c]]
        if not gaps:
            continue
        existing = {row[c] for c in POS if row[c]}
        negs = {row[c] for c in NEG if row[c]}
        added = []
        for name, score in cand.get(row['職缺名稱'], []):
            if len(added) >= len(gaps):
                break
            if score < THRESHOLD:
                stat['rej_score'] += 1; continue
            if name not in leaves:
                stat['rej_notleaf'] += 1; continue
            if name in existing or name in added:
                stat['rej_dup'] += 1; continue
            if name in negs:
                stat['rej_neg'] += 1; continue
            added.append(name)
        for col, name in zip(gaps, added):
            df.at[ri, col] = name
            filled_cells.add((ri, col))
        if added:
            stat['rows_filled'] += 1
            stat['cells_filled'] += len(added)
        if len(added) < len(gaps):
            stat['rows_still_short'] += 1

    wb = Workbook(); ws = wb.active; ws.title = '訓練回饋'
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
    for ri, row in df.iterrows():
        ws.append([row[c] for c in cols])
    ci = {c: i + 1 for i, c in enumerate(cols)}
    for ri, col in filled_cells:
        ws.cell(row=ri + 2, column=ci[col]).fill = FILL
    ws.freeze_panes = 'B2'
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = 34 if c == '職缺名稱' else 18
    wb.save(OUT)

    print(f"\n補齊統計：")
    for k, v in stat.items():
        print(f"  {k}: {v:,}")
    print(f"\n輸出 {OUT}（{len(df):,} 列，上色 {len(filled_cells):,} 格）")
    json.dump([[r, c] for r, c in sorted(filled_cells)],
              open(f"{SCRATCH}/filled_cells.json", 'w'), ensure_ascii=False)
    return 0


if __name__ == '__main__':
    sys.exit(main())
