"""
md_sync.py — Excel ⇆ Markdown 資料層雙向同步。

工作流程鐵則：**MD 是工作資料層**。分析、修改一律讀寫 MD，不回去讀 Excel；
只有使用者要「輸出」時，才由 MD 產生 Excel。

MD 檔：不合理清單_職類校正.md（每個分頁一個 ## 區段，pipe 表格）。
欄位：eNo | 廠商編號 | 廠商名稱 | 廠商狀態 | 分區 | 客服人員 | 職缺名稱 | 現有1 | 建議1 | 現有2 | 現有3 | 現有4 | 現有5
  - 「建議1」有值 ⟺ 現有1 被判定有誤（輸出 Excel 時：現有1 刪除線+淺紅、建議1 橘底）。
  - 欄內 `|` 一律以 `／` 取代避免破壞表格。

用法:
  python3 scripts/md_sync.py to-md   [來源.xlsx]   # Excel → MD（重建資料層）
  python3 scripts/md_sync.py to-xlsx [輸出.xlsx]   # MD → Excel（要輸出時才跑）
預設來源 = 不合理清單_職類校正_對照版.xlsx；預設 MD = 不合理清單_職類校正.md
"""
import sys, warnings; warnings.filterwarnings('ignore')
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

MD   = '不合理清單_職類校正.md'
XLSX = '不合理清單_職類校正_對照版.xlsx'
COLS = ['eNo', '廠商編號', '廠商名稱', '廠商狀態', '分區', '客服人員', '職缺名稱',
        '現有1', '建議1', '現有2', '現有3', '現有4', '現有5']
# 來源 Excel 欄序（含建議1 緊鄰現有1）
XLSX_ORDER = ['廠商編號', '廠商名稱', '廠商狀態', '分區', '客服人員', '職缺編號', '職缺名稱',
              '現有1', '建議1', '現有2', '現有3', '現有4', '現有5']

def esc(v):
    return '' if v is None else str(v).replace('|', '／').strip()

def to_md(xlsx=XLSX):
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    out = ['# 不合理清單 職類校正資料層\n',
           '\n> MD 為工作資料層（唯一真實來源）。分析/修改讀此檔，勿回讀 Excel。'
           '「建議1」有值＝現有1 判定有誤。\n']
    for ws in wb.worksheets:
        out.append(f'\n## {ws.title}\n\n')
        out.append('| ' + ' | '.join(COLS) + ' |\n')
        out.append('|' + '|'.join(['---'] * len(COLS)) + '|\n')
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        hdr = [str(c) if c else '' for c in rows[0]]; i = {h: hdr.index(h) for h in hdr}
        for r in rows[1:]:
            if r[i['職缺編號']] is None:
                continue
            cells = [esc(r[i['職缺編號']]), esc(r[i['廠商編號']]), esc(r[i['廠商名稱']]),
                     esc(r[i['廠商狀態']]), esc(r[i['分區']]), esc(r[i['客服人員']]),
                     esc(r[i['職缺名稱']]), esc(r[i['現有1']]), esc(r[i.get('建議1', -1)] if '建議1' in i else ''),
                     esc(r[i['現有2']]), esc(r[i['現有3']]), esc(r[i['現有4']]), esc(r[i['現有5']])]
            out.append('| ' + ' | '.join(cells) + ' |\n')
    wb.close()
    open(MD, 'w', encoding='utf-8').writelines(out)
    print(f'✅ Excel → MD：{MD}')

def parse_md():
    sheets = {}; cur = None
    for line in open(MD, encoding='utf-8'):
        line = line.rstrip('\n')
        if line.startswith('## '):
            cur = line[3:].strip(); sheets[cur] = []
        elif line.startswith('|') and cur is not None:
            parts = [p.strip() for p in line.strip().strip('|').split('|')]
            if len(parts) != len(COLS) or parts[0] in ('eNo',) or set(parts[0]) <= set('-'):
                continue
            sheets[cur].append(dict(zip(COLS, parts)))
    return sheets

def to_xlsx(xlsx=XLSX):
    sheets = parse_md()
    RED = PatternFill('solid', fgColor='FFC7CE'); ORANGE = PatternFill('solid', fgColor='FFA500')
    STRIKE = Font(strike=True)
    wb = Workbook(); wb.remove(wb.active)
    n_strike = 0
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(XLSX_ORDER)
        for d in rows:
            ws.append([d['廠商編號'], d['廠商名稱'], d['廠商狀態'], d['分區'], d['客服人員'],
                       d['eNo'], d['職缺名稱'], d['現有1'], d['建議1'],
                       d['現有2'], d['現有3'], d['現有4'], d['現有5']])
            rr = ws.max_row
            if d['建議1']:                                   # 槓必補
                ws.cell(rr, 8).font = STRIKE; ws.cell(rr, 8).fill = RED   # 現有1
                ws.cell(rr, 9).fill = ORANGE                              # 建議1
                n_strike += 1
    wb.save(xlsx)
    print(f'✅ MD → Excel：{xlsx}（現有1 標記 {n_strike} 筆）')

if __name__ == '__main__':
    cmd = sys.argv[1] if len(sys.argv) > 1 else 'to-md'
    arg = sys.argv[2] if len(sys.argv) > 2 else None
    if cmd == 'to-md':
        to_md(arg or XLSX)
    elif cmd == 'to-xlsx':
        to_xlsx(arg or XLSX)
    else:
        print(__doc__)
