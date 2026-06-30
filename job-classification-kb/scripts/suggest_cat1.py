"""
suggest_cat1.py — 只評估「職類1（現有1）是否正確」，採 R7 職稱拆解邏輯。

依據 KB logic/02-verified-logic.md R7「職稱拆解」：複合職稱拆成多個語意成分，
每個成分對應的職類都該保護。本器只處理職類1：
  1. 把職缺名稱拆成關鍵字，對照到一組職類（KWMAP）。
  2. 現有1 只要能對應到任一拆出關鍵字的職類 → 過關（不槓）。
  3. 只有當現有1 對不到任何拆出關鍵字、且職稱有「明確（非籠統）功能職類」時，
     才槓現有1，並一定補上建議（建議=該明確功能職類，絕不留空）。
寧鬆勿濫：對不到明確功能 → 不動（擦邊保留）。

用法:
  python3 scripts/suggest_cat1.py            # 印彙整 + 寫 cat1_suggestions.json
  python3 scripts/suggest_cat1.py --md       # 另存 cat1_suggestions.md
依賴: TCode_Export.xlsx, 不合理清單_職類校正.xlsx（皆在 KB 根目錄）
"""
import sys, re, json, warnings; warnings.filterwarnings('ignore')
import openpyxl
from collections import Counter

KB_XLSX = '不合理清單_職類校正.xlsx'
TCODE   = 'TCode_Export.xlsx'

def valid_leaves():
    wb = openpyxl.load_workbook(TCODE, read_only=True); ws = wb['tCodeDutyNM']
    rows = list(ws.iter_rows(min_row=1, values_only=True)); wb.close()
    hdr = [str(c) if c else '' for c in rows[0]]; idx = {h: i for i, h in enumerate(hdr)}
    return {r[idx['CodeNameA']] for r in rows[1:] if str(r[idx['CodeType']]) == '3'}

# 關鍵字 → 合法葉名（最具體者排前）。前段＝明確功能職類；後段（GENERIC 標記）＝籠統辦公室職類，
# 只用來讓現有1「過關」，不單獨拿來當建議。
KWMAP = [
    (r'主管司機|隨扈|隨身司機',                         '主管駕駛'),
    (r'產品經理|遊戲營運\s*PM|營運\s*PM|產品\s*PM|(?<![A-Za-z])PM(?![A-Za-z])', '產品經理'),
    (r'專案經理|專案副理|專案管理師',                    '專案經理'),
    (r'儲備幹部|管理儲備|儲備主管|儲備店長|儲備',        '儲備幹部'),
    (r'業務助理|業務行政助理|業助',                      '業務助理'),
    (r'國外業務|外銷業務|貿易業務|海外業務',            '國外業務人員'),
    (r'業務主管|業務經理|業務副理|業務襄理',            '國內業務主管'),
    (r'業務專員|業務代表|業務人員|跑業務|門市業務|電話行銷|電銷', '國內業務人員'),
    (r'客服|客戶服務|客戶關係',                          '客服人員'),
    (r'會計主管|財務主管|財會主管',                      '財務／會計主管'),
    (r'會計|記帳|出納',                                  '會計／出納／記帳人員'),
    (r'採購|資材',                                       '採購人員'),
    (r'倉管|倉儲|物料管理|庫管',                          '倉管／物料管理員'),
    (r'保全|警衛|駐衛',                                  '保全人員／警衛'),
    (r'清潔|環保清潔|環境清潔|清潔隊',                   '清潔／資源回收人員'),
    (r'居家服務|居服|照服|照顧服務|照顧員',              '照顧服務／指導員'),
    (r'護理師|護士',                                     '護理師'),
    (r'教育訓練|教育訓練專員|訓練專員',                  '教育訓練人員'),
    # —— 以下為籠統辦公室職類（GENERIC），僅供現有1 過關判斷 ——
    (r'秘書|祕書|特助|特別助理',                          '秘書'),
    (r'總機|接待|櫃台|櫃檯',                              '總機／接待／櫃檯人員'),
    (r'人資|人事|人力資源|招募|徵才|獵才',                '人事／人力資源專員'),
    (r'總務',                                            '總務人員'),
    (r'工讀',                                            '工讀生'),
    (r'行政助理',                                        '行政助理'),
    (r'行政|內勤|文書|庶務',                              '行政人員'),
]
# 籠統職類：不單獨拿來建議（避免籠統→籠統的無意義替換）
GENERIC_CATS = {'秘書', '總機／接待／櫃檯人員', '人事／人力資源專員', '總務人員',
                '工讀生', '行政助理', '行政人員'}
GENERIC_PARTS = {'人員', '專員', '助理', '主管', '幹部'}

def decompose(title, VALID):
    """R7：把職缺名稱拆成關鍵字，回傳命中的 (職類, 是否籠統) 清單（依 KWMAP 具體→籠統順序）。"""
    t = (title or '').replace('／', '/')
    hits = []
    for pat, leaf in KWMAP:
        if leaf in VALID and re.search(pat, t):
            hits.append((leaf, leaf in GENERIC_CATS))
    return hits

def cur1_matches_title(cur1, title):
    """現有1 字面是否出現在職缺名稱（最直接的對上）。"""
    t = (title or '').replace('／', '/').replace('台', '檯')
    c = (cur1 or '').replace('／', '/').replace('台', '檯')
    if not c:
        return False
    if c in t:
        return True
    for part in c.split('/'):
        if len(part) >= 2 and part not in GENERIC_PARTS and part in t:
            return True
    return False

def suggest(title, cur1, VALID):
    """回傳建議改的職類；若現有1 過關或無明確功能 → None。"""
    hits = decompose(title, VALID)
    cats = {c for c, _ in hits}
    # 過關條件：現有1 字面對上 職缺名稱，或 現有1 對應到任一拆出關鍵字職類
    if cur1_matches_title(cur1, title) or cur1 in cats:
        return None
    # 只在有「明確（非籠統）功能職類」時才槓現有1，並以它為建議
    for c, is_generic in hits:
        if not is_generic:
            return c
    return None

def main():
    VALID = valid_leaves()
    wb = openpyxl.load_workbook(KB_XLSX, read_only=True); ws = wb['不合理清單']
    data = list(ws.iter_rows(min_row=2, values_only=True)); wb.close()
    out = []
    for ridx, r in enumerate(data, start=2):  # excel row number
        eno, title, cur1 = r[5], r[6], (r[7] or '')
        s = suggest(title, cur1, VALID)
        if s and s != cur1:
            out.append({'row': ridx, 'eNo': eno, 'title': title, 'cur1': cur1, 'suggest1': s})
    json.dump(out, open('cat1_suggestions.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
    # aggregate report only (token-min)
    print(f'評估列數: {len(data)}　|　現有1 建議修改: {len(out)}')
    print('\nTop 現有1 → 建議1 變更類型:')
    for (a, b), n in Counter((o['cur1'], o['suggest1']) for o in out).most_common(20):
        print(f'  {n:>4}  {a}  →  {b}')
    if '--md' in sys.argv:
        with open('cat1_suggestions.md', 'w', encoding='utf-8') as f:
            f.write('# 職類1 建議修改清單\n\n| 職缺編號 | 職缺名稱 | 現有1 | 建議1 |\n|---|---|---|---|\n')
            for o in out:
                f.write(f"| {o['eNo']} | {o['title']} | {o['cur1']} | {o['suggest1']} |\n")
        print('→ cat1_suggestions.md')

if __name__ == '__main__':
    main()
