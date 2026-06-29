"""
不合理清單_職類校正 MD <-> Excel 雙向同步工具
用法:
  python3 sync_md.py export   # Excel → MD (初始化或重建)
  python3 sync_md.py apply    # MD → Excel (套用前顯示 diff)
  python3 sync_md.py diff     # 只顯示 diff，不寫入
"""
import sys, re, warnings, pickle, pandas as pd, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

MD_PATH  = '/mnt/user-data/outputs/不合理清單_職類校正.md'
XLSX_SRC = '/mnt/user-data/outputs/不合理清單_職類校正.xlsx'   # main file (T col source)
CSV_SRC  = '/mnt/user-data/uploads/管理幕僚_人資_行政_-_不合理清單.csv'
OUT_PATH = '/mnt/user-data/outputs/不合理清單_職類校正_對照版.xlsx'
PKL_PATH = '/home/claude/duty_all.pkl'

OC = ['職務小類1','職務小類2','職務小類3','職務小類4','職務小類5']
FIX = {'護理助理':'護理助理(護佐)'}
OFFICE_MID = {'管理幕僚','人力資源','行政後勤／總務','財務會計','稽核審計','業務推廣',
 '客服開發','行銷／廣告','產品／企劃','國際貿易','法務專利','出版翻譯','新聞媒體','顧問諮詢','專案管理'}
SUPPORT = {'行政助理','業務助理','秘書','特別助理','總機／接待／櫃檯人員','行政人員',
 '人事助理','國貿助理','採購助理','行銷企劃助理','設計助理','工程助理','財務／會計助理','助教','內勤業務'}
DOMAIN = [
 (r'不動產|房屋|房仲|地產|住商|信義房屋|永慶|台灣房屋|中信房',['不動產經紀人／營業員']),
 (r'保全|警衛',['保全人員／警衛']),
 (r'人壽|保險|產險|金控|銀行',['保險業務員／經紀人','核保／保險內勤人員','理賠人員','理財專員','金融營業人員','銀行辦事人員']),
 (r'醫院|診所|醫療|長照|護理|醫學|療養|生技|藥局|照護|長期照顧|居家|居服|照服|日照|安養|養護',['護理師','護理助理(護佐)','診所助理','其他醫療從業人員','個案管理師','照顧服務／指導員','照顧實務指導員','居家服務督導員','社會工作人員／督導','職能治療師','復健技術師','就業服務員','藥師','長照機構業務負責人']),
 (r'補習|文理|教育|學校|幼兒|幼教|安親|才藝|課輔',['安親課輔老師','補習班招生人員','其他教育類老師','學校行政人員','助教','教保員／保育員','講師','補習班主任／主管']),
 (r'人力|人事|仲介|顧問|派遣|萬寶華|才庫|藝珂',['人力／外勞仲介','人事／人力資源專員','人事／人力資源主管','就業服務員','教育訓練人員']),
 (r'餐飲|餐廳|食品|飲料|茶|咖啡|烘焙|烤肉|火鍋|燒肉|餐|廚|小吃|手搖|便當|料理',['餐飲服務人員','冷熱飲調製人員','其他料理廚師','餐飲領班','餐飲主管','西點蛋糕師','麵包師','餐廚助手']),
 (r'工廠|製造|生產|產線|科技|機械|電子|金屬|塑膠|工業|實業|精密|電機',['生管／助理','生管主管','製程主管','領班／管理幹部','品管人員','包裝員／作業員','倉管／物料管理員','工廠高階主管']),
 (r'門市|超商|賣場|百貨|連鎖|超市|藥妝|門店',['門市／店員／專櫃人員','店長／門市管理人員','售票／收銀人員']),
 (r'主管司機|主管駕駛|隨扈司機|董事長.{0,3}司機|總經理.{0,3}司機',['特別助理','主管駕駛','小客車／小貨車司機']),
]

def load_duty():
    d=pickle.load(open(PKL_PATH,'rb'))
    return {x[1]:x[2] for x in d['tCodeDutyNM'] if x[4]==3}

AGENCY_PAT=re.compile(r'人力|人事|仲介|顧問|派遣|萬寶華|才庫|藝珂|人力資源|企管顧問|管理顧問')
def domain_ok(e,ctx,title=''):
    # 派遣/人力仲介/顧問公司：公司行業≠職缺本身，不以「人資domain」保護人資類項目
    # （除非職稱本身就是人資職務）
    is_agency = bool(AGENCY_PAT.search(str(ctx)))
    for pat,cats in DOMAIN:
        if re.search(pat,ctx):
            # 跳過 agency 公司的人資 domain（避免客戶職缺被誤判為人資）
            if is_agency and pat.startswith('人力|人事|仲介|顧問'):
                # 僅當職稱本身含人資詞才保護
                if not re.search(r'人資|人事|人力資源|HR|招募|徵才|教育訓練', str(title)):
                    continue
            for cat in cats:
                if cat == e or (cat in e and len(cat) > 3): return True
    return False

def lang_extra(title):
    """職稱反映語言能力需求(雙語/翻譯/口譯/語+職務) → 回傳應補的翻譯職類清單。
    排除『X文補習班』(教學科目)、純括號標註等情況。"""
    t=str(title)
    LANG={
     r'印尼語|印尼文|印尼雙語|印尼/|/印尼|Indonesian':'印尼翻譯／口譯人員',
     r'越南語|越南文|越南雙語|越南/|/越南|Vietnamese':'越南翻譯／口譯人員',
     r'泰語|泰文|泰國雙語|Thai':'泰文翻譯／口譯人員',
     r'日文|日語|日本語|Japanese':'日文翻譯／口譯人員',
     r'韓文|韓語|Korean':'韓文翻譯／口譯人員',
     r'德文|德語|German':'德文翻譯／口譯人員',
     r'法文|法語|French':'法文翻譯／口譯人員',
    }
    # 語言是「工作能力需求」的訊號：雙語/翻譯/口譯，或 語言+業務/人資/行銷/客服/秘書
    NEED=re.compile(r'雙語|翻譯|口譯|Translator|Interpreter')
    out=[]
    for pat,cat in LANG.items():
        if not re.search(pat,t): continue
        # 排除「X文/語 補習班/教室/家教/老師」= 教學科目
        if re.search(r'(印尼|越南|泰|日|韓|德|法|英)(語|文).{0,4}(補習|教室|家教|老師|班|講師)', t): continue
        # 排除純括號標註 (日文)/(英文) 且無雙語/翻譯訊號
        if re.search(r'[（(](印尼|越南|泰|日|韓|德|法|英)(語|文)[）)]', t) and not re.search(r'雙語|翻譯|口譯', t): continue
        # 需要是能力需求訊號，或語言緊接職務詞
        # 職稱同時含「語言」+「職務角色詞」→ 視為能力需求
        has_role = re.search(r'業務|人資|人事|人力資源|行銷|客服|秘書|助理|電話行銷|專員|翻譯|口譯|駐廠|店員|門市', t)
        if NEED.search(t) or has_role:
            out.append(cat)
    return out

def seniority(v):
    v=str(v)
    if re.search(r'高階|總監|協理|副總|廠長|總經理|執行長|處長',v): return 3
    if re.search(r'主管|經理|主任|襄理|副理|負責人',v): return 2
    return 1

def plan(vals, correct, ctx, title, MID):
    # 職稱明確部門標籤 → 公司 domain 不覆蓋（如「業務部門」不應因公司是派遣而保留人資項目）
    DEPT_PAT=re.compile(r'業務部[門處組]?|財務部|工務部|工程部|生產部|製造部|品保部|資訊部|總務部|研發部|行政部')
    title_has_dept=bool(DEPT_PAT.search(str(title)))
    # function tail: 職稱末段是明確的專業職務詞 → 該職務類別不被 context keyword 覆蓋
    FUNC_TAIL=re.compile(r'(人資|人事|人力資源|HR|會計|財務|稽核|法務|採購|品管|品保|工程師?|MIS|資訊)(助理|專員|主管|經理|人員|長)$')
    ft=FUNC_TAIL.search(str(title).strip())
    # 若 function tail 匹配，其對應的 tCode 中類即為優先保護對象
    FUNC_MID_MAP={'人資':'人力資源','人事':'人力資源','人力資源':'人力資源','HR':'人力資源',
                  '會計':'財務會計','財務':'財務會計','稽核':'稽核審計','法務':'法務專利',
                  '採購':'採購資材','品管':'品管品保','品保':'品管品保',
                  '工程師':'機械工程','MIS':'網路管理','資訊':'網路管理'}
    func_tail_mid=FUNC_MID_MAP.get(ft.group(1),'') if ft else ''
    title_has_dept=title_has_dept or bool(ft)  # function tail 也視同部門標籤，抑制 domain override
    correct_mid=MID.get(correct,'')
    is_office=correct_mid in OFFICE_MID
    def prot(i):
        e=vals[i]
        if not e: return False
        if e==correct: return True
        if MID.get(e)==MID.get(correct): return True
        # 職稱明確特例（如主管司機→特別助理）無條件保護
        if domain_ok(e, str(title), title): return True
        # 公司行業 domain：部門標籤抑制；非辦公室類正解需同中類
        if not title_has_dept:
            if is_office and domain_ok(e,ctx,title): return True
            if not is_office and domain_ok(e,ctx,title) and MID.get(e)==MID.get(correct): return True
        if e in str(title): return True
        # 職稱含管理詞（幹部/主管/領班/組長/課長）→ H-L 的管理類職類受保護
        # 例：環保清潔「幹部」→ 基層管理幹部 應保留（即使正解是清潔人員）
        MGMT_CATS={'基層管理幹部','領班／管理幹部','經營管理主管','儲備幹部'}
        if e in MGMT_CATS and re.search(r'幹部|主管|領班|組長|課長|股長|店長|經理|主任', str(title)):
            return True
        # 語言需求保護：職稱提到特定語言 → 對應翻譯職類不刪除
        LANG_PROTECT = {
            r'印尼語?|印尼文': '印尼翻譯／口譯人員',
            r'越南語?|越南文': '越南翻譯／口譯人員',
            r'日文|日語': '日文翻譯／口譯人員',
            r'英文|英語': '英文翻譯／口譯人員',
            r'泰文|泰語': '泰文翻譯／口譯人員',
            r'韓文|韓語': '韓文翻譯／口譯人員',
            r'德文|德語': '德文翻譯／口譯人員',
            r'法文|法語': '法文翻譯／口譯人員',
        }
        for lang_pat, trans_cat in LANG_PROTECT.items():
            if e == trans_cat and re.search(lang_pat, str(title)): return True
        # SUPPORT 保護收窄：只有「泛用行政」無條件保護；其他 SUPPORT 需與正解同中類
        BROAD_SUPPORT={'行政助理','行政人員','總機／接待／櫃檯人員','工讀生','電腦操作／資料輸入'}
        if is_office and e in BROAD_SUPPORT: return True
        if is_office and e in SUPPORT and MID.get(e)==MID.get(correct): return True
        return False
    idxs=[i for i in range(5) if vals[i]]
    title_idx=[i for i in idxs if vals[i] in str(title)]
    def redundant(i):
        if is_office and vals[i] in SUPPORT:
            for j in title_idx:
                if j!=i and MID.get(vals[i])==MID.get(vals[j]): return True
        return False
    correct_present = correct in [v for v in vals if v]
    cand=[i for i in idxs if (not prot(i)) or redundant(i)]
    # keep_strike: 正解已在、但仍有非保護的錯誤項目要槓掉
    if correct_present:
        if cand:
            cand.sort(key=lambda i:(-seniority(vals[i]),i))
            return ('keep_strike', cand)
        return ('keep', None)
    if cand:
        protected=[i for i in idxs if i not in cand]
        # 全部都是候選（無保護項目）→ 全槓，正解放第一格
        if not protected:
            return ('replace_all', cand)
        sc=seniority(correct)
        cand.sort(key=lambda i:(-(1 if seniority(vals[i])>sc else 0),-seniority(vals[i]),i))
        primary=cand[0]
        # 額外槓除：候選中「與職稱領域明顯不相關」的（非管理詞、非domain、非title提及）
        # 與 primary 不同：這些是「既非保護、又非職缺領域沾邊」的雜項
        def loosely_related(i):
            e=vals[i]
            if e in str(title): return True
            # 用完整 ctx（含公司名，如長照機構）判斷領域沾邊
            if domain_ok(e,ctx,title): return True
            if MID.get(e)==MID.get(correct): return True
            # 正解是辦公室類時：一般行政/支援項目(行政後勤、管理幕僚中類)算沾邊可共存
            if is_office and MID.get(e) in ('行政後勤／總務','管理幕僚'): return True
            # 機構型公司(長照/醫療/學校/飯店等)：行政後勤/管理類項目算機構運作沾邊，可共存
            if re.search(r'機構|醫院|診所|長照|學校|飯店|集團|中心', str(ctx)) and MID.get(e) in ('行政後勤／總務','管理幕僚'): return True
            return False
        extra_strike=[i for i in cand if i!=primary and not loosely_related(i)]
        strike_list=[primary]+extra_strike
        strike_list.sort()
        if len(strike_list)>1:
            # 把 primary 放 list 開頭，其餘排序
            rest=sorted([x for x in strike_list if x!=primary])
            return ('replace_multi', [primary]+rest)
        return ('replace', primary)
    empt=[i for i in range(5) if not vals[i]]
    if empt: return ('add',empt[0])
    c2=[i for i in idxs if MID.get(vals[i])!=MID.get(correct)] or idxs
    c2.sort(key=lambda i:(-seniority(vals[i]),i)); return ('replace',c2[0])

def extract_correct(t):
    m=re.search(r'應改為：([^（(]+)',str(t)) or re.search(r'應以「([^」]+)」',str(t))
    return m.group(1).strip() if m else ''

# ─── EXPORT: Excel/CSV → MD ─────────────────────────────────────────────────
def cmd_export():
    csv=pd.read_csv(CSV_SRC)
    wb=openpyxl.load_workbook(XLSX_SRC); ws=wb['不合理清單']
    mh=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]; midx={h:i+1 for i,h in enumerate(mh)}
    lines=['# 不合理清單 職類校正資料\n',
           '<!-- 欄位: eNo | 職缺名稱 | 廠商名稱 | 現有1~5 | 正解 | S摘要 -->\n',
           '| eNo | 職缺名稱 | 廠商名稱 | 現有1 | 現有2 | 現有3 | 現有4 | 現有5 | 正解 | S |\n',
           '|-----|---------|---------|------|------|------|------|------|-----|---|\n']
    n=0
    for r in range(2,ws.max_row+1):
        eno=ws.cell(r,midx['職缺編號']).value
        if eno is None: continue
        crow=csv[csv['職缺編號']==eno]
        if crow.empty: continue
        crow=crow.iloc[0]
        hl=[(str(crow[c]).strip() if pd.notna(crow[c]) else '') for c in OC]
        t_val=ws.cell(r,midx['企劃備註']).value
        s_val=ws.cell(r,midx['資科備註']).value
        correct=extract_correct(t_val); correct=FIX.get(correct,correct)
        # S: only keep if 不合理 text (strip long)
        s_str=str(s_val).strip() if pd.notna(s_val) and str(s_val).strip() not in ('0','nan','') else ''
        def esc(v): return str(v).replace('|','／') if v else ''
        line=f"| {eno} | {esc(crow['職缺名稱'])} | {esc(crow['廠商名稱'])} | {esc(hl[0])} | {esc(hl[1])} | {esc(hl[2])} | {esc(hl[3])} | {esc(hl[4])} | {esc(correct)} | {esc(s_str[:40])} |\n"
        lines.append(line); n+=1
    with open(MD_PATH,'w',encoding='utf-8') as f: f.writelines(lines)
    print(f"Exported {n} rows → {MD_PATH}")

# ─── PARSE MD ───────────────────────────────────────────────────────────────
def parse_md():
    rows=[]
    with open(MD_PATH,'r',encoding='utf-8') as f:
        for line in f:
            if not line.startswith('|') or line.startswith('|--'): continue
            parts=[p.strip() for p in line.strip().strip('|').split('|')]
            if len(parts)<9 or not parts[0].isdigit(): continue
            rows.append({'eNo':int(parts[0]),'職缺名稱':parts[1],'廠商名稱':parts[2],
                         '現有1':parts[3],'現有2':parts[4],'現有3':parts[5],
                         '現有4':parts[6],'現有5':parts[7],'正解':parts[8],'S':parts[9] if len(parts)>9 else ''})
    return rows

# ─── BUILD OUTPUT EXCEL ──────────────────────────────────────────────────────
def build_excel(rows, MID, do_write=True):
    csv=pd.read_csv(CSV_SRC)
    ctx_by={int(r['職缺編號']):f"{r['廠商名稱']} {r['職缺名稱']}" for _,r in csv.iterrows()}
    # load main for front/tail values
    wb_src=openpyxl.load_workbook(XLSX_SRC); ws_src=wb_src['不合理清單']
    sh=[ws_src.cell(1,c).value for c in range(1,ws_src.max_column+1)]; si={h:i+1 for i,h in enumerate(sh)}
    src_by={}
    for r in range(2,ws_src.max_row+1):
        e=ws_src.cell(r,si['職缺編號']).value
        if e: src_by[int(e)]=r

    front=['廠商編號','廠商名稱','廠商狀態','分區','客服人員','職缺編號','職缺名稱']; tail=['不合理','資科備註','企劃備註']
    cols=front[:]
    for i in range(1,6): cols+=[f'現有{i}',f'建議{i}']
    cols+=tail
    out=openpyxl.Workbook(); ws=out.active; ws.title='不合理清單'
    hf=Font(name='Arial',bold=True)
    CUR=PatternFill('solid',start_color='FFE2EFDA',end_color='FFE2EFDA')
    SUG=PatternFill('solid',start_color='FFDDEBF7',end_color='FFDDEBF7')
    DEF=PatternFill('solid',start_color='FFDAEEF3',end_color='FFDAEEF3')
    RED=PatternFill('solid',start_color='FFFFC7CE',end_color='FFFFC7CE')
    ORG=PatternFill('solid',start_color='FFFFA500',end_color='FFFFA500')
    STK=Font(strike=True)
    for nc,h in enumerate(cols,1):
        c=ws.cell(1,nc); c.value=h; c.font=hf; c.alignment=Alignment(horizontal='center')
        c.fill=CUR if h.startswith('現有') else SUG if h.startswith('建議') else DEF
    ci={h:i+1 for i,h in enumerate(cols)}
    nk=na=nr=0; outr=1; diffs=[]
    for row in rows:
        outr+=1; e=row['eNo']; correct=FIX.get(row['正解'],row['正解'])
        vals=[row[f'現有{i}'] for i in range(1,6)]
        ctx=ctx_by.get(e,''); title=row['職缺名稱']
        # copy front/tail from source
        sr=src_by.get(e)
        if sr:
            for h in front+tail:
                ws.cell(outr,ci[h]).value=ws_src.cell(sr,si[h]).value
        mode,idx=plan(vals,correct,ctx,title,MID)
        if mode in ('keep','keep_strike'): nk+=1
        elif mode=='add': na+=1
        elif mode in ('replace','replace_multi','replace_all'): nr+=1
        else: nr+=1
        for i in range(5):
            cc=ci[f'現有{i+1}']; sc=ci[f'建議{i+1}']
            ws.cell(outr,cc).value=vals[i] if vals[i] else None
            strike_slots=idx if isinstance(idx,list) else []
            if mode=='keep_strike' and i in strike_slots:
                ws.cell(outr,cc).font=STK; ws.cell(outr,cc).fill=RED
                diffs.append((e,title,f"槓除現有{i+1}「{vals[i]}」（正解已在其他格）"))
            elif mode=='replace_multi' and i in strike_slots:
                ws.cell(outr,cc).font=STK; ws.cell(outr,cc).fill=RED
                if i==strike_slots[0]:  # primary
                    ws.cell(outr,sc).value=correct; ws.cell(outr,sc).fill=ORG
                    diffs.append((e,title,f"取代現有{i+1}「{vals[i]}」→「{correct}」"))
                else:
                    diffs.append((e,title,f"槓除現有{i+1}「{vals[i]}」（與職缺領域不符）"))
            elif mode=='replace_all':
                if vals[i]:
                    ws.cell(outr,cc).font=STK; ws.cell(outr,cc).fill=RED
                if i==0:
                    ws.cell(outr,sc).value=correct; ws.cell(outr,sc).fill=ORG
                    diffs.append((e,title,f"全換：取代現有1~5全槓→「{correct}」"))
            elif idx is not None and not isinstance(idx,list) and i==idx:
                if mode=='replace':
                    ws.cell(outr,cc).font=STK; ws.cell(outr,cc).fill=RED
                    ws.cell(outr,sc).value=correct; ws.cell(outr,sc).fill=ORG
                    diffs.append((e,title,f"取代現有{i+1}「{vals[i]}」→「{correct}」"))
                elif mode=='add':
                    ws.cell(outr,sc).value=correct; ws.cell(outr,sc).fill=ORG
                    diffs.append((e,title,f"新增建議{i+1}「{correct}」"))
        # 主動補語言職類（職稱反映語言能力需求、但翻譯職類不在現有/正解/建議中）
        extra_langs=lang_extra(title)
        if extra_langs:
            existing_sug=set()
            for i in range(5):
                cv=ws.cell(outr,ci[f'現有{i+1}']).value; sv=ws.cell(outr,ci[f'建議{i+1}']).value
                if cv: existing_sug.add(cv)
                if sv: existing_sug.add(sv)
            for L in extra_langs:
                if L in existing_sug or L==correct: continue
                for i in range(5):
                    if not ws.cell(outr,ci[f'現有{i+1}']).value and not ws.cell(outr,ci[f'建議{i+1}']).value:
                        ws.cell(outr,ci[f'建議{i+1}']).value=L
                        ws.cell(outr,ci[f'建議{i+1}']).fill=ORG
                        existing_sug.add(L)
                        diffs.append((e,title,f"補語言職類建議{i+1}「{L}」"))
                        break
    for nc,h in enumerate(cols,1):
        w=15 if (h.startswith('現有') or h.startswith('建議')) else {'廠商名稱':22,'職缺名稱':32,'資科備註':40,'企劃備註':46}.get(h,11)
        ws.column_dimensions[get_column_letter(nc)].width=w
    ws.freeze_panes='H2'
    if do_write:
        out.save(OUT_PATH); print(f"Saved {OUT_PATH}")
    print(f"不動:{nk} 新增:{na} 取代:{nr}")
    return diffs

# ─── DIFF ────────────────────────────────────────────────────────────────────
def cmd_diff():
    MID=load_duty(); rows=parse_md()
    print(f"MD rows: {len(rows)}")
    diffs=build_excel(rows,MID,do_write=False)
    print(f"\n差異筆數: {len(diffs)}")
    for e,t,d in diffs[:20]: print(f"  eNo {e} [{t[:20]}] {d}")
    if len(diffs)>20: print(f"  ...共 {len(diffs)} 筆")

# ─── APPLY ────────────────────────────────────────────────────────────────────
def cmd_apply():
    MID=load_duty(); rows=parse_md()
    print(f"MD rows: {len(rows)}")
    diffs=build_excel(rows,MID,do_write=False)
    print(f"\n=== 即將套用的差異（前30筆）===")
    for e,t,d in diffs[:30]: print(f"  eNo {e} [{t[:22]}] {d}")
    if len(diffs)>30: print(f"  ...共 {len(diffs)} 筆")
    print(f"\n確認套用後請重新執行: python3 sync_md.py apply --confirm")
    if '--confirm' in sys.argv:
        build_excel(rows,MID,do_write=True)
        print("已寫入 Excel")

if __name__=='__main__':
    cmd=sys.argv[1] if len(sys.argv)>1 else 'help'
    if cmd=='export': cmd_export()
    elif cmd=='diff': cmd_diff()
    elif cmd=='apply': cmd_apply()
    else: print(__doc__)
