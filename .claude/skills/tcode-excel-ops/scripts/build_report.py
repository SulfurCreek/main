#!/usr/bin/env python3
"""Build a multi-sheet 異動清單 xlsx (異動方式 | CodeNo | CodeNameA).

Usage:
  build_report.py <spec.json> <out.xlsx>

spec.json:
{
  "name_source": {                      # where to read CodeNameA from per sheet
    "tCodeCertify":   "/path/newest.xlsx",
    "tCodeDutyPT":    "/path/other.xlsx"
  },
  "changes": {
    "tCodeCertify":   [["ADD", [180547,180548]], ["EDIT", [183800,230135]]],
    "tCodeWorkAbility":[["ADD", [110700,140217]]]
  }
}
RENAME is folded to EDIT. ADD fill green E2EFDA, EDIT fill blue DEEAF1.
Save under /mnt/user-data/outputs/ then call present_files.
"""
import sys, json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def norm(s): return "" if s is None else str(s).strip()
def asint(v):
    try: return int(float(v))
    except: return None

def name_index(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames: return {}
    ws = wb[sheet]
    h = {norm(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if norm(ws.cell(1, c).value)}
    no_c = h.get("CodeNo", 4); na_c = h.get("CodeNameA", 5)
    return {asint(ws.cell(r, no_c).value): norm(ws.cell(r, na_c).value)
            for r in range(2, ws.max_row + 1) if asint(ws.cell(r, no_c).value)}

JH = "Microsoft JhengHei"
HF = Font(name=JH, bold=True, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="4472C4")
AF = Font(name=JH); SEPF = Font(name=JH, bold=True, color="595959")
FILLS = {"ADD": PatternFill("solid", fgColor="E2EFDA"),
         "EDIT": PatternFill("solid", fgColor="DEEAF1")}

def fold(ct):  # RENAME -> EDIT
    ct = ct.upper()
    return "EDIT" if ct in ("RENAME", "MOVE_EDIT") else ct

def build(spec, out_path):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for sheet, groups in spec["changes"].items():
        src = spec["name_source"].get(sheet)
        idx = name_index(src, sheet) if src else {}
        ws = wb.create_sheet(sheet)
        ws.append(["異動方式", "CodeNo", "CodeNameA"])
        for c in range(1, 4):
            cell = ws.cell(1, c); cell.font = HF; cell.fill = HFILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        flat = [(fold(ct), code) for ct, codes in groups for code in codes]
        cur = None
        for ct, code in flat:
            if ct != cur:
                ws.append([f"── {ct} ──", "", ""]); sr = ws.max_row
                ws.cell(sr, 1).font = SEPF
                ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=3)
                cur = ct
            ws.append([ct, code, idx.get(code, f"(未找到 {code})")]); rr = ws.max_row
            for c in range(1, 4):
                ws.cell(rr, c).font = AF
                if ct in FILLS: ws.cell(rr, c).fill = FILLS[ct]
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 55
    wb.save(out_path)
    print("Saved:", out_path)

if __name__ == "__main__":
    spec = json.load(open(sys.argv[1]))
    build(spec, sys.argv[2])
